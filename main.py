import openpyxl
from xlcalculator import ModelCompiler, Evaluator
import re

def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=False)

def save_workbook(wb, path):
    wb.save(path)

def update_cell(sheet, cell_ref, new_value):
    sheet[cell_ref].value = new_value

def fill_empty_cells_with_zero(sheet):
    print("🧹 Filling empty cells with 0...")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.value = 0

def extract_cell_references(formula):
    return re.findall(r'\b[A-Z]{1,3}[0-9]{1,7}\b', formula)

def validate_dependencies(sheet, formula, cell_ref, depth=0, visited=None):
    if visited is None:
        visited = set()
    indent = "  " * depth
    refs = extract_cell_references(formula)
    issues = []

    print(f"{indent}🔍 Dependencies for {cell_ref}: {refs}")
    for ref in refs:
        if ref in visited:
            print(f"{indent}⚠️ Circular reference detected at {ref}")
            continue
        visited.add(ref)

        val = sheet[ref].value
        if val is None:
            issues.append(f"{indent}⚠️ {ref} is empty")
        elif isinstance(val, str) and val.startswith("="):
            print(f"{indent}↪ {ref} contains nested formula: {val}")
            nested_issues = validate_dependencies(sheet, val[1:], ref, depth + 1, visited)
            issues.extend(nested_issues)
        else:
            print(f"{indent}✅ {ref} = {val}")

    return issues

def evaluate_excel_formula(file_path, cell_ref, sheet_name="Sheet1"):
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(file_path)
    evaluator = Evaluator(model)
    return evaluator.evaluate(f"{sheet_name}!{cell_ref}")

def evaluate_range(file_path, sheet, sheet_name, cell_range):
    print(f"\n📊 Evaluating range: {cell_range}")
    for cell_ref in cell_range:
        formula = sheet[cell_ref].value
        print(f"\n📌 Cell {cell_ref}:")
        if isinstance(formula, str) and formula.startswith("="):
            print(f"Formula: {formula}")
            issues = validate_dependencies(sheet, formula[1:], cell_ref)
            if issues:
                print("🚨 Issues detected:")
                for issue in issues:
                    print(issue)
            else:
                print("✅ All dependencies look good.")
        try:
            result = evaluate_excel_formula(file_path, cell_ref, sheet_name)
            print(f"✅ Computed value of {cell_ref}: {result}")
        except Exception as e:
            print(f"❌ Error evaluating {cell_ref}: {e}")

def main():
    file_path = "/home/vedant/Desktop/vs_code-project_folder/Suryacore2/2experiment.xlsx"
    wb = load_workbook(file_path)
    sheet = wb.active
    sheet_name = wb.sheetnames[0]

    # ✅ Fill all empty cells with 0
    fill_empty_cells_with_zero(sheet)

    # ✅ Update a cell value (example: B2 = 100)
    update_cell(sheet, "B2", 100)
    save_workbook(wb, file_path)

    # ✅ Define ranges
    f1_to_f22 = [f"F{i}" for i in range(1, 23)]
    f24_to_f41 = [f"F{i}" for i in range(24, 42)]
    h24_to_h39 = [f"H{i}" for i in range(24, 40)]

    # ✅ Evaluate all ranges
    evaluate_range(file_path, sheet, sheet_name, f1_to_f22)
    evaluate_range(file_path, sheet, sheet_name, f24_to_f41)
    evaluate_range(file_path, sheet, sheet_name, h24_to_h39)

if __name__ == "__main__":
    main()
