import openpyxl
from xlcalculator import ModelCompiler, Evaluator
from io import BytesIO
import re

def load_workbook(file_bytes):
    return openpyxl.load_workbook(file_bytes, data_only=False)

def fill_empty_cells_with_zero(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.value = 0

def update_ingredient_quantities(sheet, quantities):
    for i, qty in enumerate(quantities, start=2):
        sheet[f"B{i}"].value = qty

def extract_labels(sheet, cell_refs):
    labels = []
    for ref in cell_refs:
        col = "E" if ref.startswith("F") else "G"
        row = re.findall(r'\d+', ref)[0]
        label = sheet[f"{col}{row}"].value
        labels.append(label if label else ref)
    return labels

def is_unsupported_formula(formula):
    unsupported_keywords = ["INDEX", "OFFSET", "INDIRECT"]
    return any(keyword in formula.upper() for keyword in unsupported_keywords)

def try_manual_sumproduct(sheet, formula):
    if "SUMPRODUCT" in formula and "INDEX(AE43:AE82" in formula and "$B$2:$B$41" in formula:
        b_vals = [sheet[f"B{i}"].value or 0 for i in range(2, 42)]
        ae_vals = [sheet[f"AE{i}"].value or 0 for i in range(43, 83)]
        return sum(b * ae for b, ae in zip(b_vals, ae_vals))
    return None

def evaluate_excel_formula(file_bytes, cell_ref, sheet_name="S"):
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(file_bytes)
    evaluator = Evaluator(model)
    return evaluator.evaluate(f"{sheet_name}!{cell_ref}")

def evaluate_cells(file_bytes, quantities, cell_refs):
    wb = load_workbook(file_bytes)
    sheet = wb.active
    sheet_name = "S"

    fill_empty_cells_with_zero(sheet)
    update_ingredient_quantities(sheet, quantities)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    results = {}
    for ref in cell_refs:
        formula = sheet[ref].value
        try:
            result = evaluate_excel_formula(output, ref, sheet_name)
            results[ref] = result
        except Exception:
            if isinstance(formula, str) and formula.startswith("=") and is_unsupported_formula(formula):
                manual_result = try_manual_sumproduct(sheet, formula[1:])
                if manual_result is not None:
                    sheet[ref].value = manual_result
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    try:
                        reevaluated = evaluate_excel_formula(output, ref, sheet_name)
                        results[ref] = reevaluated
                    except Exception:
                        results[ref] = "#ERROR"
                else:
                    results[ref] = "#UNSUPPORTED"
            else:
                results[ref] = "#ERROR"
    return results, extract_labels(sheet, cell_refs)
